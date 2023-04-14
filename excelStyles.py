from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 96 Plates
frame_bg = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
frame_borders = thin_border
frame_font = Font(bold=True)
frame_alignment = Alignment(horizontal='center')

cell_bg = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
cell_borders = thin_border
cell_font = Font()
cell_alignment = Alignment(horizontal='center')


def apply_style(cell,bg=None,borders=None,font=None,alignment=None):
    if bg is not None:
        cell.fill=bg
    if borders is not None:
        cell.border=borders
    if font is not None:
        cell.font=font
    if alignment is not None:
        cell.alignment=alignment

