import openpyxl
import datetime
import warnings
from smallFunctions import *
from Plate96 import *
import xlrd

class ClientPlate:
    def __init__(self):

        # General
        self.title = None
        self.path = None


        # Associated Plates
        self.ID = None
        self.sequences = None
        self.quantities = None
        self.concentrations = None
        self.purities = None
        self.sizes = None
        self.error_rates=None
        self.flag = None
        self.MLflag = None

        # Ranges
        self.quantity_xlrange = None
        self.purity_xlrange = None
        self.error_rate_xlrange = None

        #Calculations
        self.failed_purity=0
        self.failed_quantity=0
        self.failed_flag=0
        self.failed_flag_ML=0
        self.nb_samples=0
        self.max_length=0

    def fill(self, path):

        self.path=path
        self.title = self.path.split("\\")[-1].replace(".xlsx", "")
        workbook = xlrd.open_workbook(path)

        # Oligo Tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "Oligos")
        if r:
            self.sequences=get_lines_if_sequence(values,"Sequence (5' -> 3') - Capital letters",480,0)
            for seq in self.sequences.cells:
                if seq != "":
                    self.nb_samples += 1
                    if len(seq)>self.max_length:
                        self.max_length=len(seq)
            self.sizes=get_lines_if_sequence(values,"Size",480,1)
            self.quantities=get_lines_if_sequence(values,"Quantification (pmol)",480,1)
            if self.quantities!=None:
                for q in self.quantities.cells:
                    if q<200:
                        self.failed_quantity+=1
            self.concentrations=get_lines_if_sequence(values,"Concentration (µM)",480,1)
            self.purities=get_lines_if_sequence(values,"OP2 Purity (%)",480,1)
            if self.purities!=None:
                for p in self.purities.cells:
                    if p<50:
                        self.failed_purity+=1

            self.sizes = get_lines_if_sequence(values, "Size", 480, 1)
            error_rates_comp=[(1-pow((self.purities.cells[i]/100),(1/self.sizes.cells[i])))*100 for i in range(len(self.purities.cells))]
            self.error_rates = Plate96(error_rates_comp,1)



            self.flag = get_lines_if_sequence(values, "DifficultyScore", 480, 1)
            if self.flag!=None:
                for f in self.flag.cells:
                    if f>4:
                        self.failed_flag+=1

            self.MLflag = get_lines_if_sequence(values, "ML purity prediction", 480, 1)
            if self.MLflag != None:
                for f in self.MLflag.cells:
                    if f > 4:
                        self.failed_flag_ML += 1



def get_excel_sheet_values(sheet):
    sheet_values=[]
    for row in range(1,sheet.max_row + 1):
        column=[]
        for col in range(1,sheet.max_column + 1):
            column.append(sheet.cell(row=row,column=col).value)
        sheet_values.append(column)

    return sheet_values

def get_excel_sheet_values_xlrd(sheet):
    sheet_values=[]
    for row in range(0,sheet.nrows):
        column=[]
        for col in range(0,sheet.ncols):
            column.append(sheet.cell(row,col).value)
        sheet_values.append(column)

    return sheet_values

def get_values_from_sheet_xlrd(workbook,sheetname):
    if sheetname not in workbook.sheet_names():
        print("Tab " + sheetname + " not found")
        return[0,[]]
    else:
        tab = workbook.sheet_by_name(sheetname)
        values = get_excel_sheet_values_xlrd(tab)
        return [1,values]

def f_index(sheet_values,str_looking_for):

    for row in range(len(sheet_values)):
        for col in range(len(sheet_values[0])):
            if all(word in str(sheet_values[row][col]).replace("\n"," ").replace("  "," ").split(" ") for word in str_looking_for.replace("\n"," ").replace("  "," ").split(" ")) :
                return [row,col]
    print("Couldnt find " + str_looking_for)
    return None


def next_value(sheet_values,str_looking_for):

    indexes=f_index(sheet_values,str_looking_for)
    if indexes!=None:
        return sheet_values[indexes[0]][indexes[1]+1]


def get_plate96(sheet_values,str_looking_for,is_float):

    indexes = f_index(sheet_values, str_looking_for)
    if indexes!=None:
        cells=[]
        for col in range(12):
            for row in range(8):
                value_to_add=sheet_values[indexes[0] + 1 + row][indexes[1] + 1 + col]
                if is_float:
                    if isinstance(value_to_add, datetime.datetime):
                        warnings.warn("Date found in OP2 in " + wellNbToStr(col*8+(row+1)))
                        cells.append(value_to_add.day)
                    elif value_to_add is not None and isFloat(value_to_add):
                        cells.append(float(value_to_add))
                    else:
                        cells.append(np.nan)

                else:
                    cells.append(value_to_add)
        return Plate96(cells,is_float)

def get_lines96(sheet_values,str_looking_for,is_float):

    indexes = f_index(sheet_values, str_looking_for)
    if indexes!=None:
        cells=[]
        for row in range(96):
            value_to_add=sheet_values[indexes[0] + 1 + row][indexes[1]]
            if is_float:
                if value_to_add is not None and isFloat(value_to_add):
                    cells.append(float(value_to_add))
                else:
                    cells.append(np.nan)

            else:
                cells.append(value_to_add)
        return Plate96(cells,is_float)

def get_lines_if_sequence(sheet_values,str_looking_for,lineNb,is_float):

    indexes = f_index(sheet_values, str_looking_for)
    sequence_indexes = f_index(sheet_values, "Sequence (5' -> 3') - Capital letters")
    if indexes!=None and sequence_indexes!=None:
        cells=[]
        for row in range(lineNb):
            if indexes[0]+1+row<len(sheet_values):
                if sheet_values[sequence_indexes[0]+1+row][sequence_indexes[1]] != "":
                    value_to_add=sheet_values[indexes[0] + 1 + row][indexes[1]]
                    if is_float:
                        if value_to_add is not None and isFloat(value_to_add):
                            cells.append(float(value_to_add))
                        else:
                            cells.append(np.nan)

                    else:
                        cells.append(value_to_add)
        return Plate96(cells,is_float)

def isFloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


if __name__ == '__main__':

    path = "200504_HL_P1_ncovid19_Pool1_131.xlsm"
    #workbook = openpyxl.load_workbook(path, data_only=True)
    #general_tab = workbook["General"]
    #values=get_excel_sheet_values(general_tab)


    testPlate=ClientPlate()
    testPlate.fill(path)
    print(testPlate.purities.cells)
    print(testPlate.purities.mean)
    print(testPlate.purities.cv)
    print(testPlate.purities.min)
    print(testPlate.purities.max)
    print(testPlate.flag.cells)