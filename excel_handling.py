import openpyxl
from smallFunctions import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.chart import BarChart, Series, Reference
from excelStyles import *
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as F
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from Plate96 import *

EXCEL_OUTPUT_PATH="output_excel_file.xlsm"

plate_frame_fill = PatternFill


def clear_file(wb:openpyxl.Workbook):
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])


def add96Plate(worksheet,plate96,init_row,init_column,title):

    # Create the frame
    for col in range(12):
        worksheet.cell(row=init_row,column=init_column+col+1).value=col+1
        apply_style(worksheet.cell(row=init_row,column=init_column+col+1),frame_bg,frame_borders,frame_font,frame_alignment)
    for row in range(8):
        worksheet.cell(row=init_row+row+1,column=init_column).value=chr(ord("A") + row)
        apply_style(worksheet.cell(row=init_row+row+1,column=init_column), frame_bg, frame_borders,frame_font,frame_alignment)

    worksheet.cell(row=init_row,column=init_column).value=title
    apply_style(worksheet.cell(row=init_row,column=init_column), frame_bg, frame_borders,frame_font,frame_alignment)

    # Fill the frame
    for col in range(12):
        for row in range(8):
            worksheet.cell(row=init_row + row + 1, column=init_column+col+1).value = plate96.cells[col*8+row]
            worksheet.cell(row=init_row + row + 1, column=init_column+col+1).number_format = "0.0"
            apply_style(worksheet.cell(row=init_row + row + 1, column=init_column+col+1),cell_bg,cell_borders,cell_font,cell_alignment)

    first_cell = excelColumn(init_column + 1) + str(init_row + 1)
    last_cell = excelColumn(init_column + 1 + 11) + str(init_row + 1 + 7)

    if plate96.is_float:
        #worksheet.conditional_formatting.add(first_cell+":"+last_cell,ColorScaleRule(start_type='min', start_color='FA0100',end_type='max', end_color='00AF4F'))
        if title=='Quantity':
            worksheet.conditional_formatting.add(first_cell+":"+last_cell,ColorScaleRule(start_type='num',start_value=0, start_color='FA0100',end_type='num',end_value=300, end_color='00AF4F'))
        if title=='Purity':
            worksheet.conditional_formatting.add(first_cell + ":" + last_cell,ColorScaleRule(start_type='num', start_value=0, start_color='FA0100',end_type='num', end_value=70, end_color='00AF4F'))
    #Add metadata

    worksheet.cell(row=init_row+10, column=init_column+1).value='Mean'
    worksheet.cell(row=init_row+10, column=init_column+2).value='=AVERAGE('+ first_cell + ":" + last_cell + ")"
    worksheet.cell(row=init_row+10, column=init_column+2).number_format = "0.0"

    worksheet.cell(row=init_row + 10, column=init_column + 4).value = 'CV'
    worksheet.cell(row=init_row + 10, column=init_column + 5).value = '=100*STDEV(' + first_cell + ':' + last_cell + ')/AVERAGE('+ first_cell + ':' + last_cell + ')'
    worksheet.cell(row=init_row + 10, column=init_column + 5).number_format = "0.0"

    worksheet.cell(row=init_row+10, column=init_column+7).value='Min'
    worksheet.cell(row=init_row+10, column=init_column+8).value='=MIN('+ first_cell + ":" + last_cell + ")"
    worksheet.cell(row=init_row+10, column=init_column+8).number_format = "0.0"

    worksheet.cell(row=init_row + 10, column=init_column + 10).value = 'Max'
    worksheet.cell(row=init_row + 10, column=init_column + 11).value = '=MAX(' + first_cell + ":" + last_cell + ")"
    worksheet.cell(row=init_row + 10, column=init_column + 11).number_format = "0.0"

def addHistogram(worksheet,bins,init_row,init_column,first_data_cell,last_data_cell,title,strBeforeCells):
    # Create bins
    worksheet.cell(row=init_row,column=init_column).value="Bins"
    worksheet.cell(row=init_row, column=init_column).font=Font(bold=True)

    worksheet.cell(row=init_row, column=init_column+1).value = "Range"
    worksheet.cell(row=init_row, column=init_column+1).font = Font(bold=True)

    worksheet.cell(row=init_row, column=init_column + 2).value = "Count"
    worksheet.cell(row=init_row, column=init_column + 2).font = Font(bold=True)

    offset=1
        #We separate first and last bins
    worksheet.cell(row=init_row + offset, column=init_column).value = bins[0]
    worksheet.cell(row=init_row + offset, column=init_column + 1).value = "<" + str(bins[0])
    offset=2
    for bin in bins[1:]:
        worksheet.cell(row=init_row+offset, column=init_column).value = bin
        worksheet.cell(row=init_row+offset, column=init_column+1).value = "=" + excelColumn(init_column)+ str(init_row+offset-1) + " & \"-\" & " + excelColumn(init_column)+ str(init_row+offset)
        offset=offset+1

    worksheet.cell(row=init_row + offset, column=init_column + 1).value = ">" + str(bins[-1])

    # Add Frequency function
    first_count_cell=excelColumn(init_column) + str(init_row+1)
    last_count_cell=excelColumn(init_column) + str(init_row+offset-1)
    worksheet.cell(row=init_row + 1, column=init_column + 2).value = "=FREQUENCY(" + strBeforeCells + first_data_cell + ':' + strBeforeCells + last_data_cell + ',' + first_count_cell + ':' + last_count_cell + ')'
    worksheet.formula_attributes[xlCell(init_row + 1,init_column + 2)]= {'t': 'array', 'ref': xlCell(init_row + 1,init_column + 2)+ ":" +xlCell(init_row + offset,init_column + 2)}

    # Add Chart
    barChart = BarChart()
    barChart.type="col"
    barChart.style = 2
    barChart.height = 6
    barChart.width = 13
    #barChart.title = title
    barChart.y_axis.title = 'Well count'
    barChart.x_axis.title = title + ' range'
    cp = CharacterProperties(latin=F(typeface='Calibri'), sz=800)
    barChart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    barChart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    barChart.legend=None
    data = Reference(worksheet, min_col=init_column+2, min_row=init_row+2, max_col=init_column+2, max_row=init_row+offset)
    cats = Reference(worksheet, min_col=init_column+1, min_row=init_row+2, max_col=init_column+1, max_row=init_row+offset)

    barChart.add_data(data)
    barChart.set_categories(cats)
    #barChart.shape = 3

    # Change bar filling and line color
    s = barChart.series[0]
    s.graphicalProperties.line.solidFill="4f81bd"
    s.graphicalProperties.solidFill = "4f81bd"

    barChart.x_axis.majorGridlines=None
    #barChart.y_axis.majorGridlines=None
    #barChart.plot_area.line.graphicalProperties=GraphicalProperties(solidFill="000000")
    barChart.dataLabels = DataLabelList()
    barChart.dataLabels.showVal = True

    worksheet.add_chart(barChart,xlCell(init_row,init_column+4))

if __name__ == '__main__':
    EXCEL_OUTPUT_PATH = "output_excel_file.xlsm"
    workbook=openpyxl.load_workbook(EXCEL_OUTPUT_PATH,read_only=False,keep_vba=True)
    clear_file(workbook)
    worksheet=workbook.create_sheet("Testing",0)
    add96Plate(worksheet,Plate96(list(range(96)),1),3,1,'Quantity')
    addHistogram(worksheet,[0,10,20,30,40,50,60,70,80,90],3,15,"B4","M11","Purity")
    workbook.save(filename=EXCEL_OUTPUT_PATH)