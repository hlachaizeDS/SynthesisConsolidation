import openpyxl
import datetime
import warnings
import xlrd
from smallFunctions import *
from Plate96 import *

class SynthesisPlate:
    def __init__(self):

        # General
        self.title = None
        self.path = None
        self.date = None

        # Resin
        self.resin_batch = None
        self.resin_quantity = None

        # Reagents
        self.EB_batch = None
        self.caco_batch = None
        self.Hepes_batch = None
        self.coCl2_batch = None
        self.t20_batch = None
        self.deblock_batch = None
        self.deblock_version = None
        self.wash_batch = None
        self.enzyme_batch = None
        self.enzyme_concentration = None
        self.A_batch = None
        self.C_batch = None
        self.G_batch = None
        self.T_batch = None
        self.nucs_concentration = None
        self.PSP = None
        self.Endo_b = None
        self.Endo_c = None
        self.NGS = None

        # Associated Plates
        self.ID = None
        self.sequences = None
        self.quantities = None
        self.volumes = None
        self.Volumes_all = None
        self.concentrations = None
        self.purities = None
        self.flag = None
        self.MLflag = None

        # Data recovery
        self.Concentrations_D = None
        self.Min_Concentrations_D = None
        self.Max_Concentrations_D = None
        self.CV_Concentrations_D = None
        self.Quantities_D = None
        self.Min_quantities_D = None
        self.Max_quantities_D = None
        self.CV_D = None
        self.m100_D = None
        self.m200_D = None
        self.m600_D = None
        self.Q100_D = None
        self.Q200_D = None
        self.Q600_D = None
        self.C5_D = None
        self.C8_D = None
        self.C10_D = None
        self.Volume_D = None
        self.Min_Volume_D = None
        self.Max_Volume_D = None
        self.CV_Volume_D = None

        #Calculations
        self.failed_purity=0
        self.failed_quantity=0
        self.failed_quantity2=0
        self.failed_Volume_all=0
        self.failed_flag=0
        self.failed_flag_ML=0
        self.nb_samples=0
        self.max_length=0
        self.less_than_19=0
        self.bet_20_and_59=0
        self.bet_60_and_99=0
        self.above_99=0

    def fill(self, path):

        self.path=path

        workbook=xlrd.open_workbook(path)
        self.title = self.path.split("\\")[-1].replace(".xlsx","").replace(".xlsm","")

        # General Tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "General")
        if r:
            #self.title=next_value(values,"Plate ID")
            self.date=next_value(values,"Date")

            self.resin_batch=next_value(values,"Resin batch")
            self.resin_quantity=next_value(values,"Resin qtty")

            self.EB_batch=next_value(values,"EB 2.5X")
            self.caco_batch=next_value(values,"Caco 2M")
            self.Hepes_batch=next_value(values,"HEPES")
            self.coCl2_batch=next_value(values,"CoCl2")
            self.t20_batch=next_value(values,"Tween20 10%")
            self.deblock_batch=next_value(values,"Deblock buffer")
            self.deblock_version=next_next_value(values,"Deblock buffer")
            self.wash_batch=next_value(values,"Wash buffer")

            self.enzyme_batch=next_value(values,"Enzyme")
            self.enzyme_concentration=next_value(values,"Enzyme final concentration (µM)")
            self.nucs_concentration=next_value(values,"Nucleotide final concentration (µM)")
            self.A_batch=next_value(values,"dATP-ONH2")
            self.C_batch=next_value(values,"dCTP-ONH2")
            self.G_batch=next_value(values,"dGTP-ONH2")
            self.T_batch=next_value(values,"dTTP-ONH2")

            self.PSP=next_value(values,"Cleavage type")
            self.Endo_b=next_value(values,"EndoV batch")
            self.Endo_c=next_next_value(values,"EndoV batch")
            self.NGS=next_value(values,"NGS run")

            self.Concentrations_D=next_value(values,"UV Average concentration (µM)")
            self.Min_Concentrations_D=next_next_next_value(values,"UV Min DNA quantity (pmol)")
            self.Max_Concentrations_D=next_next_next_value(values,"UV Max DNA quantity (pmol)")
            self.CV_Concentrations_D=next_next_next_value(values,"UV CV")
            self.Quantities_D=next_value(values,"UV Average DNA quantity /well (pmol)")
            self.Min_quantities_D=next_value(values,"UV Min DNA quantity (pmol)")
            self.Max_quantities_D=next_value(values,"UV Max DNA quantity (pmol)")
            self.CV_D=next_value(values,"UV CV")
            self.m100_D=next_value(values,"< 100 pmol")
            self.m200_D=next_value(values,"< 200 pmol")
            self.m600_D=next_value(values,"< 600 pmol")
            self.C5_D=next_value(values,"< 5uM")
            self.C8_D=next_value(values,"< 8uM")
            self.C10_D=next_value(values,"< 10uM")

            self.Volume_D=next_next_value(values, "UV Average DNA quantity /well (pmol)")
            self.Min_Volume_D=next_next_value(values, "UV Min DNA quantity (pmol)")
            self.Max_Volume_D=next_next_value(values, "UV Max DNA quantity (pmol)")
            self.CV_Volume_D=next_next_value(values, "UV CV")


        # Syntheses tab
        [r,values]=get_values_from_sheet_xlrd(workbook,"Syntheses")
        if r:
            self.sequences=get_plate96(values, "#Sequences", 0)
            if self.sequences==None:
                self.sequences = get_plate96(values, "Sequences", 0)

            if self.sequences != None:
                for seq in self.sequences.cells:
                    if seq!="":
                        self.nb_samples+=1
                        if len(seq)>self.max_length:
                            self.max_length=len(seq)
                        if len(seq)<20:
                            self.less_than_19 += 1
                        #elif len(seq)<40:
                            #self.bet_20_and_39 += 1
                        elif len(seq)<60:
                            self.bet_20_and_59 += 1
                        elif len(seq) < 100:
                            self.bet_60_and_99 += 1
                        #else:
                            #self.above_39 += 1
                        else:
                            self.above_99 += 1

            self.ID=get_plate96(values, "Plate Plan", 0)

        # UV Quantif tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "UV quantification")
        if r:
            self.quantities = get_plate96(values, "n (pmol)", 1)
            self.concentrations = get_plate96(values, "C (µM)", 1)
            if self.quantities!=None:
                for q in self.quantities.cells:
                    if q<200:
                        self.failed_quantity+=1
                    if q<100:
                       self.failed_quantity2+=1

        self.Volume_Ratio = next_value(values, "Ratio Vol. Real/Calc.")
        self.Real_volume = next_value(values, "Vol. Real")

        # UV quantification tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "UV quantification")
        if r:
            self.Volumes_all = get_plate96(values, "V = (µL)", 1)
            if self.Volumes_all != None:
                for v in self.Volumes_all.cells:
                    if v < 50:
                        self.failed_Volume_all += 1


       # OP2 tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "OP2")
        if r:
            self.purities = get_plate96(values, "Purity", 1)
            if self.purities==None:
                self.purities = get_plate96(values, "Purity (N)", 1)

            if self.purities!=None:
                for p in self.purities.cells:
                    if p<50:
                        self.failed_purity+=1

        #BILAN Tab
        [r, values] = get_values_from_sheet_xlrd(workbook, "BILAN")
        if r:
            self.flag = get_lines96(values,"DifficultyScore" , 1)
            if self.flag!=None:
                for f in self.flag.cells:
                    if f > 4:
                        self.failed_flag+=1

            self.MLflag = get_lines96(values, "ML predicted purity", 1)
            if self.MLflag != None:
                for f in self.MLflag.cells:
                    if f < 0.5:
                        self.failed_flag_ML += 1

def get_excel_sheet_values(sheet):
    sheet_values=[]
    for row in range(1,sheet.max_row + 1):
        column=[]
        for col in range(1,sheet.max_column + 1):
            column.append(sheet.cell(row=row,column=col).value)
        sheet_values.append(column)

    return sheet_values

def get_excel_sheet_values_xlrd(sheet):
    sheet_values=[]
    for row in range(0,sheet.nrows):
        column=[]
        for col in range(0,sheet.ncols):
            column.append(sheet.cell(row,col).value)
        sheet_values.append(column)

    return sheet_values

def get_values_from_sheet_xlrd(workbook,sheetname):
    if sheetname not in workbook.sheet_names():
        print("Tab " + sheetname + " not found")
        return[0,[]]
    else:
        tab = workbook.sheet_by_name(sheetname)
        values = get_excel_sheet_values_xlrd(tab)
        return [1,values]

# def f_index(sheet_values,str_looking_for):
#
#     for row in range(len(sheet_values)):
#         for col in range(len(sheet_values[0])):
#             if sheet_values[row][col]==str_looking_for :
#                 return [row,col]
#     print("Couldnt find " + str_looking_for)
#     return None

def f_index(sheet_values,str_looking_for):
    for row in range(len(sheet_values)):
        for col in range(len(sheet_values[0])):
            if str(sheet_values[row][col]).replace("\n"," ").replace("  "," ").split(" ") == str_looking_for.replace("\n"," ").replace("  "," ").split(" "):
                return [row,col]
    print("Couldnt find " + str_looking_for)
    return None


def next_value(sheet_values,str_looking_for):

    indexes=f_index(sheet_values,str_looking_for)
    if indexes!=None:
        try:
            return sheet_values[indexes[0]][indexes[1]+1]
        except:
            print("index out of range for " + str_looking_for)


def next_next_value(sheet_values,str_looking_for):

    indexes=f_index(sheet_values,str_looking_for)
    if indexes!=None:
        try :
            return sheet_values[indexes[0]][indexes[1]+2]
        except:
            print("index out of range for " + str_looking_for)

def next_next_next_value(sheet_values,str_looking_for):

    indexes=f_index(sheet_values,str_looking_for)
    if indexes!=None:
        try :
            return sheet_values[indexes[0]][indexes[1]+3]
        except:
            print("index out of range for " + str_looking_for)

def get_plate96(sheet_values,str_looking_for,is_float):

    indexes = f_index(sheet_values, str_looking_for)
    if indexes!=None:
        cells=[]
        for col in range(12):
            for row in range(8):
                value_to_add=sheet_values[indexes[0] + 1 + row][indexes[1] + 1 + col]
                if is_float:
                    if isinstance(value_to_add, datetime.datetime):
                        warnings.warn("Date found in OP2 in " + wellNbToStr(col*8+(row+1)))
                        cells.append(value_to_add.day)
                    elif value_to_add is not None and isFloat(value_to_add):
                        cells.append(float(value_to_add))
                    else:
                        cells.append(np.nan)

                else:
                    cells.append(value_to_add)
        return Plate96(cells,is_float)

def get_lines96(sheet_values,str_looking_for,is_float):

    indexes = f_index(sheet_values, str_looking_for)
    if indexes!=None:
        cells=[]
        for row in range(96):
            value_to_add=sheet_values[indexes[0] + 1 + row][indexes[1]]
            if is_float:
                if value_to_add is not None and isFloat(value_to_add):
                    cells.append(float(value_to_add))
                else:
                    cells.append(np.nan)

            else:
                cells.append(value_to_add)
        return Plate96(cells,is_float)

def isFloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


if __name__ == '__main__':

    path = "200504_HL_P1_ncovid19_Pool1_131.xlsm"
    #workbook = openpyxl.load_workbook(path, data_only=True)
    #general_tab = workbook["General"]
    #values=get_excel_sheet_values(general_tab)


    testPlate=SynthesisPlate()
    testPlate.fill(path)
    print(testPlate.purities.cells)
    print(testPlate.purities.mean)
    print(testPlate.purities.cv)
    print(testPlate.purities.min)
    print(testPlate.purities.max)
    print(testPlate.flag.cells)